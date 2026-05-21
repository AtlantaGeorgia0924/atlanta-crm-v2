#!/usr/bin/env python3
"""Incremental service import from INVENTORY-3.xlsx into service_jobs.

Hardened behavior:
- non-destructive upsert by legacy_source_id
- phone-first client dedupe
- placeholder row skipping
- validated date parsing with warnings
- structured logging and retry logic
- sync error logging into sync_errors when available
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
import time
import uuid
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

from app.db.supabase_client import get_supabase  # noqa: E402


load_dotenv(os.path.join(os.path.dirname(__file__), "..", "backend", ".env"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("import_services_from_excel")

BATCH_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", "200"))
RETRIES = int(os.getenv("IMPORT_RETRIES", "4"))


def log_event(event: str, **fields):
    logger.info(json.dumps({"event": event, **fields}, default=str))


def retry_call(fn, operation: str):
    last_exc = None
    for attempt in range(RETRIES):
        try:
            return fn()
        except Exception as exc:
            last_exc = exc
            if attempt == RETRIES - 1:
                break
            time.sleep(0.5 * (attempt + 1))
            log_event("retry", operation=operation, attempt=attempt + 1, error=str(exc))
    raise RuntimeError(f"{operation} failed after retries: {last_exc}")


def is_placeholder(value) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    return text in {".", "..", "...", "....", ",,"}


def normalize_phone(value) -> str | None:
    digits = re.sub(r"\D+", "", str(value or ""))
    return digits or None


def parse_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        return float(text)
    except Exception:
        return default


def parse_date(value, row_num: int, field_name: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    log_event("malformed_date", row=row_num, field=field_name, raw_value=text)
    return None


def normalize_payment_status(raw_status: str, amount_charged: float, paid_amount: float, is_return: bool) -> str:
    text = str(raw_status or "").strip().upper()
    if is_return or text == "RETURNED":
        return "RETURNED"
    if text in {"PAID", "UNPAID", "PART PAYMENT", "PARTIAL"}:
        if text == "PARTIAL":
            return "PART PAYMENT"
        return text
    if paid_amount <= 0:
        return "UNPAID"
    if paid_amount < amount_charged:
        return "PART PAYMENT"
    return "PAID"


def calculate_profit(amount_charged: float, paid_amount: float, expense_amount: float, payment_status: str) -> float:
    if payment_status in {"UNPAID", "RETURNED"}:
        realized = 0.0
    elif payment_status == "PART PAYMENT":
        realized = paid_amount
    else:
        realized = paid_amount if paid_amount > 0 else amount_charged
    return realized - expense_amount


def log_sync_error(sb, legacy_source_id: str | None, operation: str, error_message: str):
    payload = {
        "table_name": "service_jobs",
        "legacy_source_id": legacy_source_id,
        "operation": operation,
        "error_message": str(error_message)[:500],
        "created_at": datetime.utcnow().isoformat(),
    }
    try:
        sb.table("sync_errors").insert(payload).execute()
    except Exception as exc:
        log_event("sync_error_log_failed", operation=operation, error=str(exc))


def upsert_batches(sb, table: str, rows: list[dict], on_conflict: str):
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        retry_call(
            lambda batch=batch: sb.table(table).upsert(batch, on_conflict=on_conflict).execute(),
            operation=f"upsert_{table}",
        )


def import_services() -> dict:
    sb = get_supabase()
    file_path = os.path.join(os.path.dirname(__file__), "..", "data", "INVENTORY-3.xlsx")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb["Sheet1"]

    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}

    existing_clients = retry_call(
        lambda: sb.table("clients").select("id,name,phone").execute().data or [],
        operation="read_existing_clients",
    )
    clients_by_phone = {normalize_phone(c.get("phone")): c for c in existing_clients if normalize_phone(c.get("phone"))}
    clients_by_name = {str(c.get("name") or "").strip().upper(): c for c in existing_clients if str(c.get("name") or "").strip()}

    services_payload: list[dict] = []
    clients_upserts: list[dict] = []
    seen_legacy_ids: set[str] = set()
    parse_errors = 0
    skipped_rows = 0
    conflicts = 0

    def get_cell(cells, col_name):
        col_idx = headers.get(col_name)
        if not col_idx or col_idx > len(cells):
            return None
        cell = cells[col_idx - 1]
        return cell.value if hasattr(cell, "value") else cell

    for row_idx in range(2, ws.max_row + 1):
        try:
            row = list(ws[row_idx])
            name = str(get_cell(row, "NAME") or "").strip()
            phone = normalize_phone(get_cell(row, "PHONE NUMBER"))
            description = str(get_cell(row, "DESCRIPTION") or "").strip()
            amount_charged = parse_float(get_cell(row, "PRICE"), 0.0)
            paid_amount = parse_float(get_cell(row, "Amount paid"), 0.0)
            expense_amount = parse_float(get_cell(row, "EXPENSE AMOUNT"), 0.0)
            paid_date = parse_date(get_cell(row, "PAID DATE"), row_idx, "PAID DATE")
            service_date = parse_date(get_cell(row, "DATE"), row_idx, "DATE")
            raw_status = str(get_cell(row, "STATUS") or "").strip()
            record_id = str(get_cell(row, "RECORD_ID") or "").strip()

            if is_placeholder(name) or is_placeholder(description) or amount_charged <= 0:
                skipped_rows += 1
                continue

            if record_id:
                legacy_source_id = f"excel_service:{record_id}"
            else:
                base = f"{name}|{description}|{service_date or ''}|{amount_charged}|{row_idx}"
                legacy_source_id = f"excel_service:{uuid.uuid5(uuid.NAMESPACE_DNS, base)}"

            if legacy_source_id in seen_legacy_ids:
                conflicts += 1
                log_sync_error(sb, legacy_source_id, "duplicate_payload_key", "Duplicate legacy_source_id in file payload")
                continue
            seen_legacy_ids.add(legacy_source_id)

            is_return = raw_status.upper() == "RETURNED"
            payment_status = normalize_payment_status(raw_status, amount_charged, paid_amount, is_return)
            if payment_status == "RETURNED":
                paid_amount = 0.0

            # Phone-first client dedupe.
            matched_client = clients_by_phone.get(phone) if phone else None
            if not matched_client:
                matched_client = clients_by_name.get(name.upper())

            if matched_client:
                client_id = matched_client.get("id")
                client_name = matched_client.get("name") or name
            else:
                client_id = str(uuid.uuid4())
                client_name = name
                new_client = {
                    "id": client_id,
                    "legacy_source_id": f"excel_client:{phone or row_idx}",
                    "name": name,
                    "phone": phone,
                }
                clients_upserts.append(new_client)
                if phone:
                    clients_by_phone[phone] = new_client
                clients_by_name[name.upper()] = new_client

            services_payload.append(
                {
                    "legacy_source_id": legacy_source_id,
                    "client_id": client_id,
                    "client_name": client_name,
                    "service_name": description[:100] if description else "Service",
                    "description": description,
                    "quantity": 1.0,
                    "amount_charged": amount_charged,
                    "paid_amount": paid_amount,
                    "payment_status": payment_status,
                    "paid_date": paid_date,
                    "service_date": service_date,
                    "due_date": service_date,
                    "service_expense_amount": expense_amount,
                    "calculated_profit": calculate_profit(amount_charged, paid_amount, expense_amount, payment_status),
                    "is_return": payment_status == "RETURNED",
                }
            )
        except Exception as exc:
            parse_errors += 1
            log_event("row_parse_error", row=row_idx, error=str(exc))

    if clients_upserts:
        upsert_batches(sb, "clients", clients_upserts, "id")

    if services_payload:
        upsert_batches(sb, "service_jobs", services_payload, "legacy_source_id")

    result = {
        "rows_processed": ws.max_row - 1,
        "services_upserted": len(services_payload),
        "clients_upserted": len(clients_upserts),
        "rows_skipped": skipped_rows,
        "parse_errors": parse_errors,
        "conflicts": conflicts,
    }
    log_event("import_summary", **result)
    return result


if __name__ == "__main__":
    try:
        summary = import_services()
        print(json.dumps(summary, indent=2))
        raise SystemExit(0)
    except Exception as exc:
        logger.exception("Service import failed")
        raise SystemExit(1)
