#!/usr/bin/env python3
"""Incremental inventory import from STOCK-3.xlsx into inventory_items."""

from __future__ import annotations

import json
import logging
import os
import re
import sys
import time
import uuid
from datetime import datetime, timezone

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))
from app.db.supabase_client import get_supabase  # noqa: E402


logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("import_inventory_from_excel")

BATCH_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", "200"))
RETRIES = int(os.getenv("IMPORT_RETRIES", "4"))


def log_event(event: str, **fields):
    logger.info(json.dumps({"event": event, **fields}, default=str))


def retry_call(fn, operation: str):
    last_exc = None
    for attempt in range(RETRIES):
        try:
            return fn()
        except Exception as exc:
            last_exc = exc
            if attempt == RETRIES - 1:
                break
            time.sleep(0.5 * (attempt + 1))
            log_event("retry", operation=operation, attempt=attempt + 1, error=str(exc))
    raise RuntimeError(f"{operation} failed after retries: {last_exc}")


def parse_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace("NGN", "").replace("N", "").replace(",", "").replace("₦", "").strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def is_placeholder(value) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    return text in {".", "..", "...", "....", ",,"}


def to_iso_date(value, row_num: int, field_name: str):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc).isoformat()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%B %d, %Y", "%b %d, %Y", "%B, %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc).isoformat()
        except Exception:
            continue
    log_event("malformed_date", row=row_num, field=field_name, raw_value=s)
    return None


def normalize_status(product_status: str, sold_date_raw) -> str:
    # sold date alone MUST NOT imply PAID.
    s = (product_status or "").strip().upper()
    if "RETURN" in s:
        return "RETURNED"
    if "PARTIAL" in s or "PART PAYMENT" in s:
        return "PART PAYMENT"
    if "PAID" in s or s == "SOLD":
        return "PAID"
    if "UNPAID" in s or "PENDING" in s or "AVAILABLE" in s:
        return "UNPAID"
    # Unknown status defaults to UNPAID even if sold date exists.
    return "UNPAID"


def calc_product_profit(selling_price: float, cost_price: float, expense_amount: float, payment_status: str) -> float:
    if payment_status in {"UNPAID", "RETURNED"}:
        realized_revenue = 0.0
    else:
        realized_revenue = selling_price
    return realized_revenue - cost_price - expense_amount


def log_sync_error(sb, legacy_source_id: str | None, operation: str, error_message: str):
    payload = {
        "table_name": "inventory_items",
        "legacy_source_id": legacy_source_id,
        "operation": operation,
        "error_message": str(error_message)[:500],
        "created_at": datetime.utcnow().isoformat(),
    }
    try:
        sb.table("sync_errors").insert(payload).execute()
    except Exception as exc:
        log_event("sync_error_log_failed", operation=operation, error=str(exc))


def main():
    root = os.path.join(os.path.dirname(__file__), "..")
    load_dotenv(os.path.join(root, "backend", ".env"))

    file_path = os.path.join(root, "data", "STOCK-3.xlsx")
    if not os.path.exists(file_path):
        logger.error("Missing Excel file: %s", file_path)
        return 1

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb["Sheet1"]
    headers = {str(c.value): i for i, c in enumerate(ws[3], start=1) if c.value}

    sb = get_supabase()
    rows_to_upsert = []
    rows_skipped = 0
    parse_errors = 0
    conflicts = 0
    seen_legacy_ids: set[str] = set()

    now_iso = datetime.now(timezone.utc).isoformat()

    def get_cell(cells, key):
        idx = headers.get(key)
        if not idx or idx > len(cells):
            return None
        cell = cells[idx - 1]
        return cell.value if hasattr(cell, "value") else cell

    for r in range(4, ws.max_row + 1):
        try:
            cells = list(ws[r])
            description = str(get_cell(cells, "DESCRIPTION") or "").strip()
            device = str(get_cell(cells, "DEVICE") or "").strip()
            color = str(get_cell(cells, "COLOUR") or "").strip()
            storage = str(get_cell(cells, "STORAGE") or "").strip()
            imei = str(get_cell(cells, "IMEI") or "").strip()
            record_id = str(get_cell(cells, "RECORD_ID") or "").strip()
            seller = str(get_cell(cells, "NAME OF SELLER") or "").strip()
            product_status = str(get_cell(cells, "PRODUCT STATUS") or "").strip()
            sold_date_raw = get_cell(cells, "AVAILABILITY/DATE SOLD")
            expense_amount = parse_float(get_cell(cells, "EXPENSE AMOUNT"))
            expense_desc = str(get_cell(cells, "EXPENSE DESCRIPTION") or "").strip()
            cost_price = parse_float(get_cell(cells, "COST PRICE"))

            if is_placeholder(product_status) or (is_placeholder(description) and is_placeholder(record_id)):
                rows_skipped += 1
                continue

            if record_id:
                legacy_source_id = f"excel_inventory:{record_id}"
            else:
                base = f"{description}|{imei}|{device}|{r}"
                legacy_source_id = f"excel_inventory:{uuid.uuid5(uuid.NAMESPACE_DNS, base)}"

            if legacy_source_id in seen_legacy_ids:
                conflicts += 1
                log_sync_error(sb, legacy_source_id, "duplicate_payload_key", "Duplicate legacy_source_id in file payload")
                continue
            seen_legacy_ids.add(legacy_source_id)

            item_name = description or f"Stock Item {record_id or r}"
            sku = imei or record_id or None
            category = device or None
            quantity = 1.0
            selling_price = cost_price
            payment_status = normalize_status(product_status, sold_date_raw)
            paid_date = to_iso_date(sold_date_raw, r, "AVAILABILITY/DATE SOLD")
            product_profit = calc_product_profit(selling_price, cost_price, expense_amount, payment_status)

            extra_parts = []
            if color:
                extra_parts.append(f"Color: {color}")
            if storage:
                extra_parts.append(f"Storage: {storage}")
            if seller:
                extra_parts.append(f"Supplier: {seller}")
            if expense_desc:
                extra_parts.append(f"Expense Description: {expense_desc}")
            full_description = item_name
            if extra_parts:
                full_description = f"{item_name} | " + " | ".join(extra_parts)

            rows_to_upsert.append(
                {
                    "legacy_source_id": legacy_source_id,
                    "item_name": item_name,
                    "sku": sku,
                    "imei": imei or None,
                    "category": category,
                    "description": full_description,
                    "quantity": quantity,
                    "unit": "pcs",
                    "cost_price": cost_price,
                    "selling_price": selling_price,
                    "expense_amount": expense_amount,
                    "product_profit": product_profit,
                    "payment_status": payment_status,
                    "paid_date": paid_date,
                    "is_return": payment_status == "RETURNED",
                    "source_created_at": now_iso,
                    "source_updated_at": now_iso,
                }
            )
        except Exception as exc:
            parse_errors += 1
            log_event("row_parse_error", row=r, error=str(exc))

    for i in range(0, len(rows_to_upsert), BATCH_SIZE):
        chunk = rows_to_upsert[i : i + BATCH_SIZE]
        retry_call(
            lambda chunk=chunk: sb.table("inventory_items").upsert(chunk, on_conflict="legacy_source_id").execute(),
            operation="upsert_inventory_items",
        )

    result = {
        "rows_processed": ws.max_row - 3,
        "rows_upserted": len(rows_to_upsert),
        "rows_skipped": rows_skipped,
        "parse_errors": parse_errors,
        "conflicts": conflicts,
    }
    log_event("import_summary", **result)
    print(json.dumps(result, indent=2))
    return 0 if parse_errors == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
